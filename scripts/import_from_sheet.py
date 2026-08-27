#!/usr/bin/env python3
"""
import_from_sheet.py — แปลง Google Sheet ของร้านให้เป็น data/*.csv ที่ระบบใช้งานได้

ต้นทาง: https://docs.google.com/spreadsheets/d/1d_dMTYFI0opKvJdgFoqRezIwjav25ZrfGjcANLD7dnc
ชีตที่ใช้:
  - "SKU Price Comparison" → รายการสินค้า
  - "Reference"            → ราคาผู้ขายพร้อมแหล่งอ้างอิง

วิธีอัปเดตข้อมูลรอบถัดไป:
  1. ดาวน์โหลดชีตเป็น .xlsx (File > Download > Microsoft Excel)
  2. วางทับ data/source/sinthai_sheet_snapshot.xlsx
  3. python3 scripts/import_from_sheet.py
  4. python3 scripts/build_catalog.py     # ดูผลกระทบต่อกำไร

สคริปต์นี้ "ไม่แก้ตัวเลขให้เงียบๆ" — ทุกอย่างที่ต้องเดาหรือแก้ จะติดธงไว้ในคอลัมน์
data_flags เสมอ เพื่อให้ตามกลับไปตรวจได้
"""

import argparse
import csv
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DEFAULT_SNAPSHOT = DATA / "source" / "sinthai_sheet_snapshot.xlsx"

# ราคาที่ติดหมายเหตุนี้คือค่าตัวอย่างที่ติดมากับเทมเพลตตั้งแต่แรก ไม่ใช่ราคาที่ไปสำรวจมาจริง
# ชีตติดสถานะว่า "Verified" ไว้ ซึ่งขัดกับหมายเหตุของตัวเอง — ห้ามเอาไปคิดเป็นต้นทุน
TEMPLATE_SAMPLE_MARKER = "original sample data retained from template"

# Makro = ที่ร้านไปซื้อของ (ต้นทุน) | Lotus, BigC = ร้านคู่แข่ง (เพดานราคาขาย)
VENDOR_ROLE = {"Makro": "supplier", "Lotus": "competitor", "BigC": "competitor"}

# แปลง Match Type ในชีตให้เป็นระดับความน่าเชื่อถือที่ระบบใช้
MATCH_TYPE_CONFIDENCE = {
    "official exact pack": "verified",
    "official product page": "verified",
    "closest comparable": "comparable",
    "closest listing": "comparable",
    "comparable listing": "comparable",
    "comparable flavor listing": "comparable",
    "official comparable flavor": "comparable",
    "derived from official single listing": "comparable",
    "official listing": "listing",
    "official search listing": "listing",
    "official category listing": "listing",
    "official/closest category page": "listing",
    "official/brand page": "listing",
    "closest/exact search listing": "listing",
    "search link": "listing",
}


def clean(value):
    """ตัดช่องว่างและทำให้อักขระไทยอยู่ในรูปแบบมาตรฐานเดียวกัน

    ชื่อแบรนด์ในชีตมีทั้ง 'ฟันโอ' และ 'ฟันโอ ' (มีช่องว่างท้าย) ซึ่งถ้าไม่ normalize
    จะกลายเป็นคนละแบรนด์ตอน group

    ตัวเลขที่ไม่มีทศนิยมต้องไม่กลายเป็น '325.0' — สำคัญมากกับบาร์โค้ด เพราะ
    openpyxl อ่านบาร์โค้ดล้วนตัวเลขมาเป็น float ทำให้ได้ '8851959142011.0'
    ซึ่งยิงสแกนแล้วจะหาไม่เจอ
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = unicodedata.normalize("NFC", text)
    return re.sub(r"\s+", " ", text).strip()


def format_number(value):
    """เขียนตัวเลขแบบไม่มี .0 ห้อยท้าย แต่ยังเก็บทศนิยมจริงไว้ (1.5 ต้องคงเป็น 1.5)"""
    if value is None:
        return ""
    return str(int(value)) if float(value).is_integer() else str(value)


def to_number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_sku(raw):
    """SKU001 และ SKU0003 ปนกันในชีต → ปรับให้เป็น SKU + เลข 4 หลักทั้งหมด"""
    text = clean(raw)
    m = re.fullmatch(r"SKU\s*0*(\d+)", text, flags=re.IGNORECASE)
    return f"SKU{int(m.group(1)):04d}" if m else None


# รูปแบบจำนวนต่อแพ็คที่พบจริงในชื่อสินค้า เรียงจากเจาะจงที่สุดไปหากว้างที่สุด
PACK_PATTERNS = [
    r"(?:แพ็ค|แพ็k|pack)\s*(\d+)",        # "แพ็ค 12", "Pack 6"
    r"[xX×]\s*(\d+)\s*(?:ชิ้น|units?|pcs?)?",  # "x 12", "x12 units"
    r"(\d+)\s*ชิ้น",                        # "12 ชิ้น"
    r"(\d+)\s*(?:ก้อน|แผง|ซอง)",            # "60ก้อน", "20แผง"
]


def parse_pack_qty(product_name):
    """ดึงจำนวนชิ้นต่อแพ็คออกจากชื่อสินค้า

    จำเป็นเพราะราคาซัพพลายเออร์บางเจ้าให้มาเป็นราคาต่อชิ้น ถ้าไม่รู้จำนวนต่อแพ็ค
    จะเทียบกับราคาขายแบบยกแพ็คไม่ได้ คืนค่า (จำนวน, เจอหรือไม่)
    """
    name = clean(product_name)
    for pattern in PACK_PATTERNS:
        matches = re.findall(pattern, name, flags=re.IGNORECASE)
        if matches:
            # ชื่อแบบ "(กล่อง15แพ็ค60ก้อน)" มีหลายตัวเลข — เอาค่ามากสุดคือจำนวนชิ้นรวม
            qty = max(int(m) for m in matches)
            if 1 < qty <= 500:
                return qty, True
    return 1, False


def load_products(workbook):
    sheet = workbook["SKU Price Comparison"]
    headers = [clean(c.value) for c in sheet[1]]
    col = {h: i for i, h in enumerate(headers) if h}

    def cell(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    products, unassigned = [], []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        name = clean(cell(row, "Product Name"))
        if not name:
            continue

        sku = normalize_sku(cell(row, "SKU Code"))
        pack_qty, pack_found = parse_pack_qty(name)
        flags = []
        if not pack_found:
            flags.append("pack_qty_not_in_name")

        unit = clean(cell(row, "Unit"))
        # "600 ml" แต่หน่วยเขียนว่า L — แก้ให้ตรงกับตัวเลขขนาดบรรจุ
        pack_size = to_number(cell(row, "Pack Size"))
        if unit.upper() == "L" and pack_size and pack_size > 10:
            unit = "ml"
            flags.append("unit_corrected_L_to_ml")

        barcode = clean(cell(row, "Barcode"))
        if not barcode:
            flags.append("barcode_missing")

        category = clean(cell(row, "Category"))
        if not category:
            flags.append("category_missing")

        retail = to_number(cell(row, "Price"))
        if retail is None:
            flags.append("no_retail_price")

        record = {
            "sku_code": sku,
            # เก็บรหัสเดิมที่เขียนไว้ในชีตด้วย เพราะพนักงานยังเรียก SKU001 อยู่
            "legacy_sku_code": clean(cell(row, "SKU Code")),
            "barcode": barcode,
            "brand": clean(cell(row, "Brand")),
            "product_name": name,
            "category": category,
            "subcategory": clean(cell(row, "Subcategory")),
            "pack_size": format_number(pack_size),
            "pack_size_unit": unit,
            "pack_qty": pack_qty,
            "retail_price": "" if retail is None else f"{retail:.2f}",
            "active": "TRUE",
            "_flags": flags,
        }
        (products if sku else unassigned).append(record)

    return products, unassigned


def assign_missing_skus(products, unassigned):
    """ออกรหัส SKU ให้สินค้าที่ยังไม่มี

    Phase 1 กำหนดว่า SKU เป็นค่าที่ต้องมี สินค้าที่ไม่มีรหัสจึงเข้าระบบไม่ได้เลย
    เรียงตาม (แบรนด์, ชื่อ) ก่อนออกรหัส เพื่อให้รันกี่ครั้งก็ได้รหัสเดิมเสมอ
    ไม่ใช่เลื่อนไปมาตามลำดับแถวในชีต
    """
    used = {p["sku_code"] for p in products}
    highest = max((int(s[3:]) for s in used), default=0)

    for record in sorted(unassigned, key=lambda r: (r["brand"], r["product_name"])):
        highest += 1
        record["sku_code"] = f"SKU{highest:04d}"
        record["_flags"].append("sku_assigned_by_import")
        products.append(record)
    return products


def load_vendor_prices(workbook, known_skus):
    sheet = workbook["Reference"]
    rows = []
    seen = set()
    skipped_unknown = 0

    for row in sheet.iter_rows(min_row=4, values_only=True):
        sku = normalize_sku(row[0])
        if not sku:
            continue
        vendor = clean(row[3])
        price = to_number(row[4])
        if price is None or vendor not in VENDOR_ROLE:
            continue  # แถว "All Vendors" ที่ยังไม่มีราคา = ตัวตั้งรอกรอก ข้ามไป
        if sku not in known_skus:
            skipped_unknown += 1
            continue

        note = clean(row[9])
        match_type = clean(row[6]).lower()
        confidence = MATCH_TYPE_CONFIDENCE.get(match_type, "listing")
        if TEMPLATE_SAMPLE_MARKER in note.lower():
            # ค่าตัวอย่างจากเทมเพลต ไม่ใช่ราคาจริง — กันไม่ให้ถูกใช้คำนวณต้นทุน
            confidence = "template_sample"

        checked = row[7]
        checked_date = checked.date().isoformat() if hasattr(checked, "date") else clean(checked)[:10]

        key = (sku, vendor, checked_date)
        if key in seen:
            continue
        seen.add(key)

        rows.append({
            "sku_code": sku,
            "vendor": vendor,
            "vendor_role": VENDOR_ROLE[vendor],
            "price": f"{price:.2f}",
            "price_basis": "pack",
            "confidence": confidence,
            "checked_date": checked_date,
            "source_url": clean(row[5]),
            "note": note,
        })

    return rows, skipped_unknown


PRODUCT_COLUMNS = [
    "sku_code", "legacy_sku_code", "barcode", "brand", "product_name",
    "category", "subcategory", "pack_size", "pack_size_unit", "pack_qty",
    "retail_price", "active", "data_flags",
]
VENDOR_COLUMNS = [
    "sku_code", "vendor", "vendor_role", "price", "price_basis",
    "confidence", "checked_date", "source_url", "note",
]


def apply_barcode_captures(products):
    """เติมบาร์โค้ดที่พนักงานยิงจากตัวสินค้าจริงผ่านหน้า "เก็บบาร์โค้ด" ในแอป

    ทำไมต้องมีไฟล์นี้แยกจากชีต: ชีตต้นทางมีบาร์โค้ดแค่ 2 จาก 141 รายการ และการยิง
    จากสินค้าจริงคือทางเดียวที่ได้บาร์โค้ดที่ถูกต้อง ถ้าไม่เก็บไว้ที่นี่ บาร์โค้ดที่
    อุตส่าห์เดินยิงทั้งร้านจะหายทันทีที่ import รอบถัดไป เพราะ products_master.csv
    ถูกสร้างใหม่จากชีตทุกครั้ง

    ลำดับความสำคัญ: ถ้าชีตมีบาร์โค้ดอยู่แล้ว ให้ชีตชนะ (ถือว่าเจ้าของร้านตั้งใจแก้)
    """
    path = DATA / "barcode_captures.csv"
    if not path.exists():
        return 0, []

    with open(path, encoding="utf-8-sig", newline="") as f:
        captures = {
            clean(row["sku_code"]): clean(row["barcode"])
            for row in csv.DictReader(f)
            if clean(row.get("sku_code")) and clean(row.get("barcode"))
        }
    if not captures:
        return 0, []

    # บาร์โค้ดเดียวห้ามผูกกับสินค้าหลายตัว ไม่งั้นยิงขายแล้วไม่รู้ว่าตัวไหน และ
    # unique index ในฐานข้อมูลจะไม่ยอมรับด้วย เจอซ้ำเมื่อไหร่ให้ "ข้ามทั้งคู่"
    # ไม่ใช่เลือกมาตัวหนึ่ง เพราะเราไม่รู้ว่าอันไหนคือของจริง
    owners = {}
    for sku, barcode in captures.items():
        owners.setdefault(barcode, []).append(sku)

    conflicts = []
    blocked = set()
    for barcode, skus in owners.items():
        if len(skus) > 1:
            conflicts.append(f"{barcode} ถูกยิงให้ทั้ง {', '.join(sorted(skus))} — ข้ามทั้งหมด ต้องยิงใหม่")
            blocked.update(skus)

    applied = 0
    for record in products:
        barcode = captures.get(record["sku_code"])
        if not barcode or record["barcode"] or record["sku_code"] in blocked:
            continue
        record["barcode"] = barcode
        record["_flags"] = [f for f in record["_flags"] if f != "barcode_missing"]
        record["_flags"].append("barcode_from_capture")
        applied += 1

    unknown = sorted(set(captures) - {p["sku_code"] for p in products})
    if unknown:
        conflicts.append(f"อ้างถึง SKU ที่ไม่มีในชีต: {', '.join(unknown)}")
    return applied, conflicts


def write_csv(path, columns, records):
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--snapshot", type=Path, default=DEFAULT_SNAPSHOT,
                        help="ไฟล์ .xlsx ที่ดาวน์โหลดจาก Google Sheet")
    args = parser.parse_args()

    if not args.snapshot.exists():
        sys.exit(f"ไม่พบไฟล์ {args.snapshot} — ดาวน์โหลดชีตเป็น .xlsx มาวางไว้ก่อน")

    workbook = openpyxl.load_workbook(args.snapshot, data_only=True)
    products, unassigned = load_products(workbook)
    products = assign_missing_skus(products, unassigned)
    products.sort(key=lambda r: r["sku_code"])

    captured, capture_warnings = apply_barcode_captures(products)

    known = {p["sku_code"] for p in products}
    vendor_prices, skipped = load_vendor_prices(workbook, known)
    vendor_prices.sort(key=lambda r: (r["sku_code"], r["vendor"]))

    priced = {v["sku_code"] for v in vendor_prices
              if v["vendor_role"] == "supplier" and v["confidence"] != "template_sample"}
    for record in products:
        if record["sku_code"] not in priced:
            record["_flags"].append("no_supplier_price")
        record["data_flags"] = ";".join(dict.fromkeys(record.pop("_flags")))

    write_csv(DATA / "products_master.csv", PRODUCT_COLUMNS, products)
    write_csv(DATA / "vendor_prices.csv", VENDOR_COLUMNS, vendor_prices)

    samples = sum(1 for v in vendor_prices if v["confidence"] == "template_sample")
    assigned = sum(1 for p in products if "sku_assigned_by_import" in p["data_flags"])
    no_barcode = sum(1 for p in products if "barcode_missing" in p["data_flags"])

    print(f"สินค้า               : {len(products)} รายการ")
    print(f"  ออกรหัส SKU ให้ใหม่ : {assigned} รายการ (เดิมไม่มีรหัส)")
    print(f"  บาร์โค้ดจากการยิงจริง: {captured} รายการ")
    print(f"  ไม่มีบาร์โค้ด        : {no_barcode} รายการ")
    for warning in capture_warnings:
        print(f"  ⚠ บาร์โค้ดที่ยิงมา: {warning}")
    print(f"ราคาผู้ขาย            : {len(vendor_prices)} แถว")
    print(f"  เป็นค่าตัวอย่างเทมเพลต: {samples} แถว (ไม่นับเป็นต้นทุน)")
    print(f"  รู้ต้นทุนจริง         : {len(priced)} รายการ")
    if skipped:
        print(f"ข้ามราคาที่อ้างถึง SKU ที่ไม่มีในชีตสินค้า: {skipped} แถว")


if __name__ == "__main__":
    main()
